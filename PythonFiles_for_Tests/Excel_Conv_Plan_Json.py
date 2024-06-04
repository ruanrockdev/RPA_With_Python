# Vamos converter xlsx in Json and more, lets do it ;)
import openpyxl
import json

wb = openpyxl.load_workbook("products.xlsx")
sheet = wb["Sheet1"]

products = {}

# Obtendo o total de linhas com dados para usar
nTotalLinhas = sheet.max_row

# Obtendo o total de colunas com dados
nTotalColunas = sheet.max_column

for row in sheet.iter_rows(min_row=2, max_row=nTotalLinhas, min_col=1, max_col=nTotalColunas, values_only=True):
    product_id = row[0]
    product = {
        "name": row[1],
        "price": row[2],
        "avaliable": row[3],
        "Color": row[4]
    }
    products[product_id] = product

print(json.dumps(products, indent=4))
# for row in sheet.iter_rows(min_row=1, max_row=nTotalLinhas, min_col=1, max_col=3, values_only=True):
#     print(row)
