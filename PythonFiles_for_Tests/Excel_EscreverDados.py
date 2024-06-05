import openpyxl

# wb = openpyxl.Workbook()
# sheet = wb["Sheet"]  # Abrir com nome Sheet por ser padrão

# # colocando um valor em uma celular metodo 1
# sheet["A1"] = "Hello, World!"

# # colocando um valor na celular metodo 2
# sheet.cell(row=2, column=2).value = "Ruan Sabe"

# # Salvando planilha
# wb.save("Nova3.xlsx")

# Método com append em tuplas
wb = openpyxl.Workbook()
sheet = wb.active

rows = (
    (88, 11, 34),
    (22, 22, 22),
    (33, 33, 33)
)

for row in rows:
    sheet.append(row)
wb.save("Append.xlsx")
