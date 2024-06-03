import openpyxl

# Abrindo a planilha
wb = openpyxl.load_workbook("Excel_PlanilhaExemplo.xlsx")
print("Mostrando o Tipo do arquivo", type(wb))
print("Mostrando o nome das abas da planilha", wb.sheetnames)

sheet = wb["Sheet1"]
print(sheet)

# Retorna qual aba está ativa no momento
activeSheet = wb.active
print(activeSheet)

# Retorna a quantidade de linhas com dados
print("Quantidade de Linhas:", sheet.max_row)
nTotalLinhas = sheet.max_row

# Retorna a quantidade de colunas com dados
print("Quantidade de Colunas:", sheet.max_column)

# Retorna valores de celulas específicas
print(sheet["A1"].value)

# Mostrar propriedades de uma celula em particular
v = sheet["A1"]
# Modo String Interpolation
print('Row %s, column %s is %s' % (v.row, v.column, v.value))
# Modo F Simplificado
print(f'Row {v.row}, column {v.column} is {v.value}')
print('Cell %s is %s' % (v.coordinate, v.value))

print("Conteúdo da célula linha 1 e coluna 2 =",
      sheet.cell(row=1, column=2).value)

# Mostrar dados de uma coluna da planilha
# for i in range(1, nTotalLinhas + 1):
#     print(i, sheet.cell(row=i, column=2).value)

# Obter um Range específico da planilha
# print(tuple(sheet["A1:C7"]))

# Obter os valores de um Range específico da planilha
# for rowOfCellObjects in sheet["A1:C7"]:
#     for cellObj in rowOfCellObjects:
#         print(cellObj.coordinate, cellObj.value)
#     print("--- END ROW ---")

# Modo mais simples de obter os valores de uma planilha Usando desempacotamento de variaveis
# Aloca pra uma variavel o RANGE das celulas, para cada coluna
print("\n")
cells = sheet["A1:C7"]
print(type(cells))
# For com as colunas nessa váriavel (loop na tabela)
print(f"Mostando valores SEM índice da celular\n")
for c1, c2, c3 in cells:
    # Mostrando apenas valores
    print(c1.value, c2.value, c3.value)

print("\n")
print(f"Mostando valores com índice da celular\n")
for c1, c2, c3 in cells:
    # Mostrando valores com índice da celula
    print(c1.coordinate, c1.value, c2.coordinate,
          c2.value, c3.coordinate, c3.value)

# mostrar dados de uma coluna, precisa converter pra uma lista antes
print(list(sheet.columns)[1])
for c in list(sheet.columns)[1]:
    print(c.value)

# Usar modos para planilha dos mais variados, inclusive
